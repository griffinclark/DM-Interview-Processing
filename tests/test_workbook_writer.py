from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from planlock.config import Settings
from planlock.models import (
    AccountCandidate,
    CanonicalPlanDocument,
    ExpenseCandidate,
    FieldCandidate,
    HoldingCandidate,
    ValueKind,
)
from planlock.workbook_writer import (
    apply_assignments_to_workbook,
    build_assignments,
    copy_locked_template,
)


def test_build_assignments_and_preserve_formula_driven_fields(tmp_path: Path) -> None:
    settings = Settings.from_env()
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    workbook_path = copy_locked_template(settings, job_dir)

    document = CanonicalPlanDocument(
        fields={
            "profile.client_1.first_name": FieldCandidate(
                target_key="profile.client_1.first_name",
                value="Taylor",
                value_kind=ValueKind.STRING,
                page_number=1,
                source_excerpt="Taylor",
                confidence=0.9,
            ),
        },
        expenses=[
            ExpenseCandidate(
                category="Travel",
                label="Imported Target",
                monthly_amount=1250.0,
                page_number=1,
                source_excerpt="Travel $1,250 suggested",
                confidence=0.92,
                comment="Suggested monthly target from planner.",
            )
        ],
        accounts=[
            AccountCandidate(
                net_worth_section="asset",
                account_type="Savings",
                owner_name="Joint",
                account_identifier="xxxx1238",
                institution="Ally Bank",
                balance=15000.25,
                monthly_contribution=300.0,
                page_number=3,
                source_excerpt="Ally Bank savings ending 1238 balance $15,000.25",
                confidence=0.91,
            ),
            AccountCandidate(
                net_worth_section="liability",
                account_type="Mortgage",
                owner_name="Joint",
                institution="HomeLoanServ",
                balance=402500.10,
                page_number=4,
                source_excerpt="Mortgage principal balance $402,500.10",
                confidence=0.89,
            ),
        ],
        holdings=[
            HoldingCandidate(
                sheet_name="Taxable Accounts",
                owner_section="client_1",
                account_name="Brokerage",
                holding_name="Vanguard Total International Stock Fund",
                symbol="VTIAX",
                shares=120.0,
                price=23.5,
                purchase_price=25.0,
                page_number=2,
                source_excerpt="Buy $6,000 of Vanguard Total International Stock Fund",
                confidence=0.8,
            )
        ],
    )

    assignments, warnings = build_assignments(document)
    assert not warnings
    apply_assignments_to_workbook(workbook_path, assignments)

    workbook = load_workbook(workbook_path)
    assert workbook["Data Input"]["C6"].value == "Taylor"
    assert workbook["Data Input"]["B18"].value == "Savings"
    assert workbook["Data Input"]["G18"].value == 15000.25
    assert workbook["Data Input"]["B19"].value == "Mortgage"
    assert workbook["Data Input"]["G19"].value == 402500.10
    assert workbook["Net Worth"]["B6"].value == "Ally Bank - Savings - Joint - xxxx1238"
    assert workbook["Net Worth"]["C6"].value == 15000.25
    assert workbook["Net Worth"]["B22"].value == "HomeLoanServ - Mortgage - Joint"
    assert workbook["Net Worth"]["C22"].value == 402500.10
    assert workbook["Expenses"]["D55"].value == 1250.0
    assert workbook["Expenses"]["C55"].value == '=IF(D55="","",D55*12)'
    assert workbook["Taxable Accounts"]["B6"].value == "Brokerage"
    assert workbook["Taxable Accounts"]["K7"].value == '=IF(AND(I7<>"",J7<>""),I7*J7,"")'
    assert workbook["Taxable Accounts"]["M7"].value == '=IF(AND(I7<>"",J7<>"",L7<>""),I7*(J7-L7),"")'


def test_copy_locked_template_preserves_prepopulated_transactions(tmp_path: Path) -> None:
    settings = Settings.from_env()
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    workbook_path = copy_locked_template(settings, job_dir)

    workbook = load_workbook(workbook_path)

    assert workbook["Transactions Raw"]["A1"].value == "account"
    assert workbook["Transactions Raw"]["A2"].value == "Amex Joint"
    assert workbook["Transactions Raw"]["F2"].value == "Aftercare/Childcare/Tuition"
