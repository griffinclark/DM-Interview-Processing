from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

import planlock.template_entry_agent as entry_agent_module
from planlock.models import (
    AgentQuestion,
    EntrySessionState,
    FieldCandidate,
    PageOcrResult,
    SheetEntryResult,
    ValueKind,
)
from planlock.template_entry_agent import (
    LangGraphTemplateEntryAgent,
    answer_from_question,
    read_sheet_context,
)


class FakeGraph:
    def __init__(self, results: list[SheetEntryResult]) -> None:
        self._results = list(results)
        self.prompts: list[str] = []

    def invoke(self, state):
        self.prompts.append(state["prompt"])
        return {
            "sheet_name": state["sheet_name"],
            "prompt": state["prompt"],
            "result": self._results.pop(0),
        }


def _session_state(tmp_path: Path) -> EntrySessionState:
    return EntrySessionState(
        job_id="job-123",
        template_sha256="template-sha",
        workbook_path=tmp_path / "output.xlsx",
        ocr_results_path=tmp_path / "ocr_results.json",
        sheet_order=["Data Input"],
    )


def _ocr_results() -> list[PageOcrResult]:
    return [
        PageOcrResult(
            page_number=1,
            summary="Client profile summary",
            raw_text="Client 1 first name: Taylor",
            source_snippets=["Client 1 first name: Taylor"],
            confidence=0.99,
        )
    ]


def _custom_ocr_results(*, raw_text: str, source_snippet: str, summary: str = "OCR summary") -> list[PageOcrResult]:
    return [
        PageOcrResult(
            page_number=1,
            summary=summary,
            raw_text=raw_text,
            source_snippets=[source_snippet],
            confidence=0.99,
        )
    ]


def _pending_question() -> AgentQuestion:
    return AgentQuestion(
        id="client-1-name",
        sheet_name="Data Input",
        prompt="Which first name should populate client 1?",
        rationale="The OCR appears to contain multiple candidate names.",
        affected_targets=["profile.client_1.first_name"],
    )


def test_advance_reuses_raw_pdf_rereview_answer_before_asking(tmp_path: Path) -> None:
    question = _pending_question()
    initial_result = SheetEntryResult(sheet_name="Data Input", question=question)
    resolved_result = SheetEntryResult(
        sheet_name="Data Input",
        mapped_fields=[
            FieldCandidate(
                target_key="profile.client_1.first_name",
                value="Taylor",
                value_kind=ValueKind.STRING,
                page_number=1,
                source_excerpt="Client 1 first name: Taylor",
                confidence=0.94,
            )
        ],
    )

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._graph = FakeGraph([initial_result, resolved_result])  # type: ignore[attr-defined]
    agent._build_prompt = lambda **kwargs: "sheet prompt"  # type: ignore[attr-defined]
    agent._review_question_against_raw_pdf = lambda **kwargs: answer_from_question(  # type: ignore[attr-defined]
        question,
        answer="Taylor",
        source="raw_pdf_review",
    )

    session_state = _session_state(tmp_path)
    updated_state = agent.advance(session_state, _ocr_results())

    assert updated_state.completed is True
    assert updated_state.pending_question is None
    assert len(updated_state.sheet_results) == 1
    assert updated_state.sheet_results[0].mapped_fields[0].value == "Taylor"
    assert len(updated_state.user_answers) == 1
    assert updated_state.user_answers[0].source == "raw_pdf_review"
    assert updated_state.user_answers[0].answer == "Taylor"
    assert updated_state.questions_asked == []
    assert len(agent._graph.prompts) == 2  # type: ignore[attr-defined]


def test_advance_marks_question_after_raw_pdf_rereview_when_no_answer_found(tmp_path: Path) -> None:
    initial_result = SheetEntryResult(sheet_name="Data Input", question=_pending_question())

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._graph = FakeGraph([initial_result])  # type: ignore[attr-defined]
    agent._build_prompt = lambda **kwargs: "sheet prompt"  # type: ignore[attr-defined]
    agent._review_question_against_raw_pdf = lambda **kwargs: None  # type: ignore[attr-defined]

    session_state = _session_state(tmp_path)
    updated_state = agent.advance(session_state, _ocr_results())

    assert updated_state.completed is False
    assert updated_state.pending_question is not None
    assert updated_state.pending_question.pdf_rereviewed is True
    assert len(updated_state.questions_asked) == 1
    assert updated_state.questions_asked[0].pdf_rereviewed is True
    assert len(agent._graph.prompts) == 1  # type: ignore[attr-defined]


def test_invoke_structured_with_retries_uses_shared_retry_helper(monkeypatch) -> None:
    observed: dict[str, object] = {}

    class FakeLLM:
        def invoke(
            self,
            *,
            schema,
            messages,
            operation_name=None,
            timeout_seconds=None,
            usage_notifier=None,
            progress_notifier=None,
            tools=None,
        ):
            observed["schema"] = schema
            observed["messages"] = messages
            observed["operation_name"] = operation_name
            observed["timeout_seconds"] = timeout_seconds
            observed["usage_notifier"] = usage_notifier
            observed["progress_notifier"] = progress_notifier
            observed["tools"] = tools
            return SheetEntryResult(sheet_name="Data Input")

    def fake_invoke_with_retries(settings, operation_name, invoke_fn, retry_notifier=None):
        observed["settings"] = settings
        observed["operation_name"] = operation_name
        observed["retry_notifier"] = retry_notifier
        return invoke_fn(90.0)

    monkeypatch.setattr(entry_agent_module, "invoke_with_retries", fake_invoke_with_retries)

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._settings = object()
    agent._llm = FakeLLM()

    notifier = object()
    result = agent._invoke_structured_with_retries(
        operation_name="Workbook entry for Data Input",
        schema=SheetEntryResult,
        messages=["message"],
        retry_notifier=notifier,
    )

    assert isinstance(result, SheetEntryResult)
    assert observed["schema"] is SheetEntryResult
    assert observed["settings"] is agent._settings
    assert observed["operation_name"] == "Workbook entry for Data Input"
    assert observed["retry_notifier"] is notifier
    assert observed["messages"] == ["message"]
    assert observed["usage_notifier"] is None
    assert observed["progress_notifier"] is None
    assert observed["tools"] is None


def test_read_sheet_context_includes_all_populated_cells(monkeypatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Input"
    for row_number in range(1, 206):
        sheet[f"A{row_number}"] = f"Value {row_number}"
    workbook.save(workbook_path)

    monkeypatch.setitem(
        entry_agent_module.ALLOWED_WRITE_CELLS_BY_SHEET,
        "Data Input",
        {f"A{row_number}" for row_number in range(1, 206)},
    )

    context = read_sheet_context(workbook_path, "Data Input")

    assert "- A1 = Value 1" in context
    assert "- A205 = Value 205" in context
    assert len(context.splitlines()) == 206


def test_build_prompt_includes_preloaded_template_data_for_non_writable_sheets(
    monkeypatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    data_input = workbook.active
    data_input.title = "Data Input"
    data_input["A1"] = "Taylor"
    transactions = workbook.create_sheet("Transactions Raw")
    transactions["A1"] = "account"
    transactions["B1"] = "date_posted"
    transactions["C1"] = "amount"
    transactions["A2"] = "Amex Joint"
    transactions["B2"] = "2024-05-02"
    transactions["C2"] = -95.70
    reference = workbook.create_sheet("Reference Data")
    reference["A1"] = "name"
    reference["B1"] = "value"
    reference["A2"] = "Primary Goal"
    reference["B2"] = "Retire at 60"
    workbook.save(workbook_path)

    monkeypatch.setitem(entry_agent_module.ALLOWED_WRITE_CELLS_BY_SHEET, "Data Input", {"A1"})

    session_state = _session_state(tmp_path)
    session_state.workbook_path = workbook_path

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._graph = FakeGraph([SheetEntryResult(sheet_name="Data Input")])  # type: ignore[attr-defined]

    updated_state = agent.advance(session_state, _ocr_results())
    prompt = agent._graph.prompts[0]  # type: ignore[attr-defined]

    assert updated_state.completed is True
    assert "Preloaded template data already present in this workbook before any agent writes:" in prompt
    assert '"sheet_name": "Reference Data"' in prompt
    assert '"data_row_count": 1' in prompt
    assert '"name": "Primary Goal"' in prompt
    assert '"value": "Retire at 60"' in prompt
    assert '"sheet_name": "Transactions Raw"' not in prompt
    assert '"account": "Amex Joint"' not in prompt


def test_data_input_sheet_starts_with_compact_ocr_context(tmp_path: Path) -> None:
    workbook_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.active.title = "Data Input"
    workbook.save(workbook_path)

    session_state = _session_state(tmp_path)
    session_state.workbook_path = workbook_path

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._graph = FakeGraph([SheetEntryResult(sheet_name="Data Input")])  # type: ignore[attr-defined]

    updated_state = agent.advance(
        session_state,
        _custom_ocr_results(
            raw_text="RAW-PDF-ONLY-TOKEN",
            source_snippet="Structured OCR snippet",
            summary="Client profile OCR summary",
        ),
    )
    prompt = agent._graph.prompts[0]  # type: ignore[attr-defined]

    assert updated_state.completed is True
    assert "Context stage: ocr_results" in prompt
    assert "Structured OCR snippet" in prompt
    assert "RAW-PDF-ONLY-TOKEN" not in prompt


def test_non_data_input_sheet_uses_data_input_before_ocr(tmp_path: Path) -> None:
    workbook_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    data_input = workbook.active
    data_input.title = "Data Input"
    data_input["C6"] = "Taylor"
    workbook.create_sheet("Expenses")
    workbook.save(workbook_path)

    session_state = EntrySessionState(
        job_id="job-123",
        template_sha256="template-sha",
        workbook_path=workbook_path,
        ocr_results_path=tmp_path / "ocr_results.json",
        sheet_order=["Expenses"],
    )

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._graph = FakeGraph(  # type: ignore[attr-defined]
        [
            SheetEntryResult(
                sheet_name="Expenses",
                unresolved_supported_targets=["expense.travel.monthly"],
            ),
            SheetEntryResult(sheet_name="Expenses"),
        ]
    )

    updated_state = agent.advance(
        session_state,
        _custom_ocr_results(
            raw_text="RAW-EXPENSES-TOKEN",
            source_snippet="Travel target $1,250",
            summary="Planner recommended travel spend",
        ),
    )

    first_prompt, second_prompt = agent._graph.prompts  # type: ignore[attr-defined]

    assert updated_state.completed is True
    assert len(agent._graph.prompts) == 2  # type: ignore[attr-defined]
    assert "Context stage: workbook_only" in first_prompt
    assert "Populated Data Input sheet context already written into the workbook:" in first_prompt
    assert "Taylor" in first_prompt
    assert "Structured OCR evidence for the document:\n- Not included in this pass." in first_prompt
    assert "RAW-EXPENSES-TOKEN" not in first_prompt
    assert "Context stage: ocr_results" in second_prompt
    assert "Travel target $1,250" in second_prompt
    assert "RAW-EXPENSES-TOKEN" not in second_prompt


def test_expenses_prompt_uses_query_transactions_tool_instead_of_sample_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    data_input = workbook.active
    data_input.title = "Data Input"
    data_input["C6"] = "Taylor"
    transactions = workbook.create_sheet("Transactions Raw")
    transactions["A1"] = "account"
    transactions["B1"] = "date_posted"
    transactions["C1"] = "amount"
    transactions["D1"] = "merchant"
    transactions["E1"] = "description"
    transactions["F1"] = "default category"
    transactions["A2"] = "Amex Joint"
    transactions["B2"] = "2024-05-02"
    transactions["C2"] = -95.70
    transactions["D2"] = "Preschool Smiles"
    transactions["E2"] = "BT*PRESCHOOL SMILES EDEN PRAIRIE        MN"
    transactions["F2"] = "Aftercare/Childcare/Tuition"
    workbook.create_sheet("Expenses")
    workbook.save(workbook_path)

    session_state = EntrySessionState(
        job_id="job-123",
        template_sha256="template-sha",
        workbook_path=workbook_path,
        ocr_results_path=tmp_path / "ocr_results.json",
        sheet_order=["Expenses"],
    )

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._graph = FakeGraph(  # type: ignore[attr-defined]
        [
            SheetEntryResult(sheet_name="Expenses", unresolved_supported_targets=["expense.travel.monthly"]),
            SheetEntryResult(sheet_name="Expenses"),
        ]
    )

    updated_state = agent.advance(session_state, _ocr_results())
    prompt = agent._graph.prompts[0]  # type: ignore[attr-defined]

    assert updated_state.completed is True
    assert "Tool available: `query_transactions`" in prompt
    assert "transactions_raw" in prompt
    assert "transactions_raw_cells" in prompt
    assert '"sheet_name": "Transactions Raw"' not in prompt
    assert '"account": "Amex Joint"' not in prompt


def test_advance_marks_preloaded_non_writable_sheet_as_template_seeded(tmp_path: Path) -> None:
    workbook_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions Raw"
    sheet["A1"] = "account"
    sheet["A2"] = "Amex Joint"
    workbook.save(workbook_path)

    session_state = EntrySessionState(
        job_id="job-123",
        template_sha256="template-sha",
        workbook_path=workbook_path,
        ocr_results_path=tmp_path / "ocr_results.json",
        sheet_order=["Transactions Raw"],
    )

    agent = object.__new__(LangGraphTemplateEntryAgent)
    updated_state = agent.advance(session_state, _ocr_results())

    assert updated_state.completed is True
    assert updated_state.sheet_summaries[0].status == "skipped"
    assert updated_state.sheet_summaries[0].message is not None
    assert "pre-populates this sheet with starter data" in updated_state.sheet_summaries[0].message


def test_advance_reorders_sheet_order_to_start_with_data_input(tmp_path: Path) -> None:
    workbook_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Input"
    workbook.create_sheet("Expenses")
    workbook.save(workbook_path)

    session_state = EntrySessionState(
        job_id="job-123",
        template_sha256="template-sha",
        workbook_path=workbook_path,
        ocr_results_path=tmp_path / "ocr_results.json",
        sheet_order=["Expenses", "Data Input"],
    )

    agent = object.__new__(LangGraphTemplateEntryAgent)
    agent._graph = FakeGraph([SheetEntryResult(sheet_name="Data Input")])  # type: ignore[attr-defined]
    agent._build_prompt = lambda **kwargs: kwargs["sheet_name"]  # type: ignore[attr-defined]
    agent._review_question_against_raw_pdf = lambda **kwargs: None  # type: ignore[attr-defined]

    updated_state = agent.advance(session_state, _ocr_results())

    assert updated_state.current_sheet_index == 1
    assert updated_state.sheet_order[0] == "Data Input"
    assert agent._graph.prompts[0] == "Data Input"  # type: ignore[attr-defined]
