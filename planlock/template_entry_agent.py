from __future__ import annotations

from datetime import date, datetime
import json
from pathlib import Path
from typing import Literal, Protocol, TypedDict

from langchain_core.messages import HumanMessage, SystemMessage
from langchain_core.tools import BaseTool
from langgraph.graph import END, StateGraph
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from planlock.config import Settings
from planlock.llm_pipeline import (
    ProgressNotifier,
    RetryNotifier,
    StructuredOutputClient,
    UsageNotifier,
    invoke_with_retries,
)
from planlock.models import (
    AgentAnswer,
    AgentQuestion,
    CellAssignment,
    CoverageSummary,
    EntrySessionState,
    PageMappingResult,
    PageOcrResult,
    QuestionOption,
    SheetEntryResult,
    SheetEntrySummary,
)
from planlock.template_schema import (
    ALLOWED_WRITE_CELLS_BY_SHEET,
    EXPENSE_ROW_BLOCKS,
    TEMPLATE_SHEET_ORDER,
    sheet_reference_for_prompt,
)
from planlock.transactions_query import (
    TRANSACTIONS_SHEET_NAME,
    build_query_transactions_tool,
    has_transaction_data,
    transactions_query_schema_reference,
)

SHEET_ENTRY_SYSTEM_PROMPT = """
You fill one locked workbook sheet at a time using OCR data, workbook state, and prior user answers.

Rules:
- Only emit candidates for the active workbook sheet.
- You may inspect the full OCR corpus for context, but only map values supported by the active sheet schema.
- If a ledger query tool is available, use it before asking the user to manually inspect or total transaction rows.
- For Net Worth, use `accounts` candidates and set `net_worth_section` to `asset` or `liability` whenever the document supports that distinction.
- Ask a user question only when a material ambiguity blocks a supported target on the active sheet.
- If the user explicitly delegates a decision back to you, choose the most defensible supported value from the available evidence and do not ask the same question again unless the sheet still cannot proceed safely.
- Emit literal values only for directly observed constants or user-provided answers.
- If a value depends on extracted constants, do not emit a computed literal. Emit the source constants and let the workbook formulas compute downstream values inside the sheet.
- Do not duplicate a derived value in both raw and computed form.
- Keep unsupported or still-ambiguous items in unresolved_supported_targets or warnings.
- Return structured output only.
""".strip()


RAW_PDF_REREVIEW_SYSTEM_PROMPT = """
You review raw OCR text before a planner-facing question is shown.

Rules:
- Use only the raw OCR text and source snippets provided in the prompt.
- Try to answer the pending question if the raw OCR contains a defensible answer.
- If the raw OCR still does not support a defensible answer, set answer_found to false.
- Do not invent facts, compute new values, or rely on information outside the prompt.
- Return structured output only.
""".strip()


class TemplateEntryAgent(Protocol):
    def advance(
        self,
        session_state: EntrySessionState,
        ocr_results: list[PageOcrResult],
        retry_notifier: RetryNotifier | None = None,
        usage_notifier: UsageNotifier | None = None,
        progress_notifier: ProgressNotifier | None = None,
    ) -> EntrySessionState:
        ...


class _GraphState(TypedDict):
    sheet_name: str
    prompt: str
    retry_notifier: RetryNotifier | None
    usage_notifier: UsageNotifier | None
    progress_notifier: ProgressNotifier | None
    tools: list[BaseTool]
    result: SheetEntryResult | None


class RawPdfQuestionReview(BaseModel):
    answer_found: bool = False
    answer: str | None = None
    rationale: str = ""
    source_page_numbers: list[int] = Field(default_factory=list)


def persist_ocr_results(path: Path, ocr_results: list[PageOcrResult]) -> None:
    path.write_text(
        json.dumps([result.model_dump(mode="json") for result in ocr_results], indent=2),
        encoding="utf-8",
    )


def load_ocr_results(path: Path) -> list[PageOcrResult]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return [PageOcrResult.model_validate(item) for item in payload]


def save_entry_state(path: Path, state: EntrySessionState) -> None:
    path.write_text(json.dumps(state.model_dump(mode="json"), indent=2), encoding="utf-8")


def load_entry_state(path: Path) -> EntrySessionState:
    return EntrySessionState.model_validate_json(path.read_text(encoding="utf-8"))


def read_sheet_context(
    workbook_path: Path,
    sheet_name: str,
    touched_cells: list[str] | None = None,
) -> str:
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[sheet_name]
    allowed_cells = sorted(ALLOWED_WRITE_CELLS_BY_SHEET.get(sheet_name, set()))
    cells_to_read = list(dict.fromkeys([*(touched_cells or []), *allowed_cells]))
    lines = [f"Workbook context for {sheet_name}:"]
    populated_lines: list[str] = []
    for cell_ref in cells_to_read:
        cell = sheet[cell_ref]
        value = cell.value
        if value in (None, ""):
            continue
        populated_lines.append(f"- {cell_ref} = {value}")
    if populated_lines:
        lines.extend(populated_lines)
    else:
        lines.append("- No writable cells are populated yet.")
    return "\n".join(lines)


def read_sheet_scaffold_context(workbook_path: Path, sheet_name: str) -> str | None:
    if sheet_name != "Expenses":
        return None

    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[sheet_name]
    lines = [
        "Locked template scaffold for Expenses (read-only reference cells, including headers and subtotal rows):"
    ]

    for row_number in range(1, 72):
        populated_cells: list[str] = []
        for column_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            value = sheet[f"{column_letter}{row_number}"].value
            if value in (None, ""):
                continue
            populated_cells.append(f"{column_letter}{row_number} = {value}")
        if populated_cells:
            lines.append(f"- row {row_number}: " + ", ".join(populated_cells))

    if all(sheet[f"A{row_number}"].value in (None, "") for row_number in range(6, 71)):
        lines.append("- Column A starter rows 6-70 are blank in the locked template.")

    if len(lines) == 1:
        lines.append("- No scaffold reference cells detected.")

    return "\n".join(lines)


def _stringify_workbook_value(value: object) -> str:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _is_formula_like_workbook_value(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _find_preloaded_table(
    workbook_path: Path,
    sheet_name: str,
    *,
    max_row_gap: int = 3,
) -> tuple[int, list[tuple[int, str]], list[tuple[int, list[object]]]] | None:
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        sheet = workbook[sheet_name]
        populated_rows: list[tuple[int, list[object]]] = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_values = list(row)
            if any(value not in (None, "") for value in row_values):
                populated_rows.append((row_number, row_values))

        for index, (row_number, row_values) in enumerate(populated_rows):
            header_pairs: list[tuple[int, str]] = []
            for column_index, value in enumerate(row_values):
                if value in (None, "") or _is_formula_like_workbook_value(value):
                    continue
                header = _stringify_workbook_value(value).strip()
                if header:
                    header_pairs.append((column_index, header))
            if len(header_pairs) < 2:
                continue

            data_rows: list[tuple[int, list[object]]] = []
            previous_row_number = row_number
            for candidate_row_number, candidate_row_values in populated_rows[index + 1 :]:
                if candidate_row_number - previous_row_number > max_row_gap:
                    break

                literal_count = sum(
                    1
                    for value in candidate_row_values
                    if value not in (None, "") and not _is_formula_like_workbook_value(value)
                )
                if literal_count >= 2:
                    data_rows.append((candidate_row_number, candidate_row_values))
                    previous_row_number = candidate_row_number

            if data_rows:
                return row_number, header_pairs, data_rows

        return None
    finally:
        workbook.close()


def read_preloaded_template_context(
    workbook_path: Path,
    *,
    sample_row_limit: int = 5,
    exclude_sheet_names: set[str] | None = None,
) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=False)
    summaries: list[dict[str, object]] = []
    excluded = {TRANSACTIONS_SHEET_NAME, *(exclude_sheet_names or set())}

    for sheet_name in workbook.sheetnames:
        if sheet_name in excluded:
            continue
        if ALLOWED_WRITE_CELLS_BY_SHEET.get(sheet_name):
            continue

        preloaded_table = _find_preloaded_table(workbook_path, sheet_name)
        if preloaded_table is None:
            continue

        header_row_number, header_pairs, data_rows = preloaded_table
        headers = [header for _, header in header_pairs]

        sample_rows: list[dict[str, object]] = []
        for row_number, row_values in data_rows[:sample_row_limit]:
            record: dict[str, str] = {}
            for index, header in header_pairs:
                if index >= len(row_values):
                    continue
                value = row_values[index]
                if value in (None, ""):
                    continue
                record[header] = _stringify_workbook_value(value)
            if record:
                sample_rows.append({"row_number": row_number, "values": record})

        summaries.append(
            {
                "sheet_name": sheet_name,
                "template_data_present": True,
                "header_row_number": header_row_number,
                "header_columns": headers,
                "data_row_count": len(data_rows),
                "sample_rows": sample_rows,
                "note": (
                    "This sheet already contains starter data from the locked template before any "
                    "agent writes. The sample rows below are illustrative; the full sheet data is "
                    "already present in the workbook."
                ),
            }
        )

    return summaries


def sheet_has_preloaded_template_data(workbook_path: Path, sheet_name: str) -> bool:
    if sheet_name == TRANSACTIONS_SHEET_NAME:
        return has_transaction_data(workbook_path)
    return _find_preloaded_table(workbook_path, sheet_name) is not None


def sheet_has_populated_writable_cells(workbook_path: Path, sheet_name: str) -> bool:
    allowed_cells = sorted(ALLOWED_WRITE_CELLS_BY_SHEET.get(sheet_name, set()))
    if not allowed_cells:
        return False
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[sheet_name]
    return any(sheet[cell_ref].value not in (None, "") for cell_ref in allowed_cells)


def prioritize_data_input_sheet(sheet_order: list[str]) -> list[str]:
    ordered = list(sheet_order)
    if "Data Input" not in ordered or not ordered or ordered[0] == "Data Input":
        return ordered
    return ["Data Input", *[sheet_name for sheet_name in ordered if sheet_name != "Data Input"]]


def touched_cells_for_assignments(assignments: list[CellAssignment], sheet_name: str) -> list[str]:
    return [assignment.cell for assignment in assignments if assignment.sheet_name == sheet_name]


def sheet_specific_prompt_guidance(sheet_name: str) -> str:
    if sheet_name == "Net Worth":
        return (
            "Net-Worth-specific guidance:\n"
            "- Use `accounts` candidates for the Family Balance Sheet rows.\n"
            "- Set `net_worth_section` to `asset` or `liability` whenever the document supports that distinction.\n"
            "- Write liability balances as positive outstanding amounts; the sheet subtracts liabilities in its formulas.\n"
            "- Keep account labels concise enough for column B while preserving institution, account type, and owner when useful.\n"
        )

    if sheet_name != "Expenses":
        return ""

    supported_categories = ", ".join(EXPENSE_ROW_BLOCKS)
    return (
        "Expenses-specific guidance:\n"
        "- Treat Expenses as a block-based summary sheet, not a fully pre-labeled line-item ledger.\n"
        f"- Supported expense categories are exactly: {supported_categories}.\n"
        "- Return at most one expense candidate per supported category block.\n"
        "- Aggregate granular Transactions Raw categories into the closest supported expense block and explain judgment calls in the candidate comment.\n"
        "- Do not ask the user to paste, screenshot, or transcribe labels from the active Expenses sheet. The workbook scaffold above is the source of truth for this tab.\n"
    )


def coverage_summary_for_state(state: EntrySessionState) -> CoverageSummary:
    supported = 0
    unresolved = 0
    unresolved_critical: list[str] = []
    critical_sheets = ["Data Input", "Expenses"]
    for result in state.sheet_results:
        mapped_count = (
            len(result.mapped_fields)
            + len(result.expenses)
            + len(result.accounts)
            + len(result.holdings)
        )
        unresolved_count = len(result.unresolved_supported_targets)
        supported += mapped_count + unresolved_count
        unresolved += unresolved_count
        if result.sheet_name in critical_sheets and mapped_count == 0 and unresolved_count > 0:
            unresolved_critical.append(result.sheet_name)
    ratio = 1.0 if supported == 0 else (supported - unresolved) / supported
    return CoverageSummary(
        supported_target_count=supported,
        unresolved_supported_target_count=unresolved,
        coverage_ratio=ratio,
        critical_sheet_names=critical_sheets,
        unresolved_critical_sheet_names=sorted(set(unresolved_critical)),
    )


def _upsert_sheet_summary(state: EntrySessionState, summary: SheetEntrySummary) -> None:
    for index, existing in enumerate(state.sheet_summaries):
        if existing.sheet_name == summary.sheet_name:
            state.sheet_summaries[index] = summary
            return
    state.sheet_summaries.append(summary)


def sheet_result_to_page_mapping_result(result: SheetEntryResult) -> PageMappingResult:
    return PageMappingResult(
        page_number=0,
        mapped_fields=result.mapped_fields,
        expenses=result.expenses,
        accounts=result.accounts,
        holdings=result.holdings,
        unmapped_items=result.unresolved_supported_targets,
        warnings=result.warnings,
    )


class LangGraphTemplateEntryAgent:
    def __init__(self, settings: Settings) -> None:
        self._settings = settings
        self._llm = StructuredOutputClient(settings, model=settings.model_mapping)
        graph = StateGraph(_GraphState)
        graph.add_node("draft_sheet", self._draft_sheet)
        graph.set_entry_point("draft_sheet")
        graph.add_edge("draft_sheet", END)
        self._graph = graph.compile()

    def _invoke_structured_with_retries(
        self,
        *,
        operation_name: str,
        schema,
        messages,
        retry_notifier: RetryNotifier | None = None,
        usage_notifier: UsageNotifier | None = None,
        progress_notifier: ProgressNotifier | None = None,
        tools: list[BaseTool] | None = None,
    ):
        invoke_kwargs = {
            "schema": schema,
            "messages": messages,
            "operation_name": operation_name,
            "tools": tools,
        }
        if usage_notifier is not None:
            invoke_kwargs["usage_notifier"] = usage_notifier
        if progress_notifier is not None:
            invoke_kwargs["progress_notifier"] = progress_notifier
        return invoke_with_retries(
            self._settings,
            operation_name,
            lambda timeout_seconds: self._llm.invoke(
                timeout_seconds=timeout_seconds,
                **invoke_kwargs,
            ),
            retry_notifier=retry_notifier,
        )

    def _draft_sheet(self, state: _GraphState) -> _GraphState:
        result = self._invoke_structured_with_retries(
            operation_name=f"Workbook entry for {state['sheet_name']}",
            schema=SheetEntryResult,
            messages=[
                SystemMessage(content=SHEET_ENTRY_SYSTEM_PROMPT),
                HumanMessage(content=state["prompt"]),
            ],
            retry_notifier=state.get("retry_notifier"),
            usage_notifier=state.get("usage_notifier"),
            progress_notifier=state.get("progress_notifier"),
            tools=state.get("tools") or None,
        )
        result.sheet_name = state["sheet_name"]
        return {
            "sheet_name": state["sheet_name"],
            "prompt": state["prompt"],
            "retry_notifier": state.get("retry_notifier"),
            "usage_notifier": state.get("usage_notifier"),
            "progress_notifier": state.get("progress_notifier"),
            "tools": state.get("tools", []),
            "result": result,
        }

    @staticmethod
    def _mapped_item_count(result: SheetEntryResult) -> int:
        return (
            len(result.mapped_fields)
            + len(result.expenses)
            + len(result.accounts)
            + len(result.holdings)
        )

    @staticmethod
    def _structured_ocr_context(ocr_results: list[PageOcrResult]) -> list[dict[str, object]]:
        payload: list[dict[str, object]] = []
        for result in ocr_results:
            payload.append(
                {
                    "page_number": result.page_number,
                    "summary": result.summary,
                    "source_snippets": result.source_snippets[:6],
                    "figures": [figure.model_dump(mode="json") for figure in result.figures],
                    "tables": [
                        {
                            "title": table.title,
                            "headers": table.headers,
                            "rows": table.rows[:6],
                        }
                        for table in result.tables
                    ],
                    "recommendations": result.recommendations,
                    "confidence": result.confidence,
                }
            )
        return payload

    @staticmethod
    def _context_stage_instructions(context_stage: Literal["workbook_only", "ocr_results"]) -> str:
        if context_stage == "workbook_only":
            return (
                "Context stage: workbook_only\n"
                "Use the active sheet, preloaded template data, and populated Data Input workbook values first. "
                "Structured OCR evidence is intentionally omitted in this pass to conserve tokens. "
                "If workbook context is insufficient, leave targets unresolved or ask a question and the runner "
                "will escalate to OCR."
            )
        return (
            "Context stage: ocr_results\n"
            "Workbook context stays primary, and structured OCR evidence is available in this pass. "
            "Use OCR only where workbook context is insufficient. Raw OCR text is still withheld unless a later "
            "raw-PDF fallback is needed."
        )

    @staticmethod
    def _latest_answers_for_sheet(
        session_state: EntrySessionState,
        *,
        sheet_name: str,
        sources: set[str],
    ) -> list[AgentAnswer]:
        latest_answers: list[AgentAnswer] = []
        seen_question_ids: set[str] = set()
        for answer in reversed(session_state.user_answers):
            if answer.sheet_name != sheet_name or answer.source not in sources:
                continue
            if answer.question_id in seen_question_ids:
                continue
            seen_question_ids.add(answer.question_id)
            latest_answers.append(answer)
        latest_answers.reverse()
        return latest_answers

    @staticmethod
    def _resume_mode_label(
        session_state: EntrySessionState,
        *,
        sheet_name: str,
    ) -> str:
        if any(answer.sheet_name == sheet_name for answer in session_state.user_answers):
            return "resumed after retaining a planner or agent answer"
        if any(summary.sheet_name == sheet_name for summary in session_state.sheet_summaries):
            return "rerunning with prior workbook writes preserved"
        return "fresh sheet pass"

    def _workflow_step_context(
        self,
        *,
        session_state: EntrySessionState,
        sheet_name: str,
    ) -> str:
        total_sheets = len(session_state.sheet_order)
        current_step_number = min(session_state.current_sheet_index + 1, total_sheets) if total_sheets else 0
        completed_sheets = [
            summary.sheet_name
            for summary in session_state.sheet_summaries
            if summary.status == "completed"
        ]
        remaining_sheets = session_state.sheet_order[session_state.current_sheet_index + 1 :]
        completed_label = ", ".join(completed_sheets) if completed_sheets else "none"
        remaining_label = ", ".join(remaining_sheets) if remaining_sheets else "none"
        return (
            "Workflow step context:\n"
            f"- Step index: {current_step_number} of {total_sheets}\n"
            f"- Active sheet: {sheet_name}\n"
            f"- Resume mode: {self._resume_mode_label(session_state, sheet_name=sheet_name)}\n"
            f"- Completed sheets so far: {completed_label}\n"
            f"- Remaining sheets after this pass: {remaining_label}\n"
            "- Prior transcript messages are intentionally dropped between passes. "
            "Use only the retained decision state below."
        )

    def _data_input_context(
        self,
        *,
        session_state: EntrySessionState,
        sheet_name: str,
    ) -> str | None:
        if sheet_name == "Data Input":
            return None
        if not sheet_has_populated_writable_cells(session_state.workbook_path, "Data Input"):
            return None
        return read_sheet_context(
            session_state.workbook_path,
            "Data Input",
            touched_cells=touched_cells_for_assignments(session_state.mapped_assignments, "Data Input"),
        )

    @staticmethod
    def _transactions_tool_available(*, workbook_path: Path, sheet_name: str) -> bool:
        return sheet_name == "Expenses" and has_transaction_data(workbook_path)

    def _sheet_tools(
        self,
        *,
        session_state: EntrySessionState,
        sheet_name: str,
    ) -> list[BaseTool]:
        if not self._transactions_tool_available(
            workbook_path=session_state.workbook_path,
            sheet_name=sheet_name,
        ):
            return []
        return [build_query_transactions_tool(session_state.workbook_path)]

    def _build_prompt(
        self,
        *,
        session_state: EntrySessionState,
        ocr_results: list[PageOcrResult],
        sheet_name: str,
        context_stage: Literal["workbook_only", "ocr_results"],
    ) -> str:
        explicit_answers_payload = [
            answer.model_dump(mode="json")
            for answer in self._latest_answers_for_sheet(
                session_state,
                sheet_name=sheet_name,
                sources={"option", "free_text"},
            )
        ]
        agent_answers_payload = [
            answer.model_dump(mode="json")
            for answer in self._latest_answers_for_sheet(
                session_state,
                sheet_name=sheet_name,
                sources={"agent", "raw_pdf_review"},
            )
        ]
        prior_assignments = [
            assignment.model_dump(mode="json")
            for assignment in session_state.mapped_assignments
            if assignment.sheet_name == sheet_name
        ]
        workbook_context = read_sheet_context(
            session_state.workbook_path,
            sheet_name,
            touched_cells=touched_cells_for_assignments(session_state.mapped_assignments, sheet_name),
        )
        scaffold_context = read_sheet_scaffold_context(session_state.workbook_path, sheet_name)
        data_input_context = self._data_input_context(
            session_state=session_state,
            sheet_name=sheet_name,
        )
        transactions_tool_available = self._transactions_tool_available(
            workbook_path=session_state.workbook_path,
            sheet_name=sheet_name,
        )
        preloaded_template_context = read_preloaded_template_context(session_state.workbook_path)
        ocr_context = self._structured_ocr_context(ocr_results) if context_stage == "ocr_results" else None
        data_input_block = ""
        if data_input_context is not None:
            data_input_block = f"Populated Data Input sheet context already written into the workbook:\n{data_input_context}\n\n"
        scaffold_block = ""
        if scaffold_context is not None:
            scaffold_block = f"{scaffold_context}\n\n"
        transactions_tool_block = ""
        if transactions_tool_available:
            transactions_tool_block = (
                "Tool available: `query_transactions`\n"
                "Use the `query_transactions` tool to inspect or aggregate the workbook's Transactions Raw "
                "ledger as a read-only SQLite database before asking the user to total transaction rows.\n"
                f"{transactions_query_schema_reference()}\n\n"
            )
        ocr_block = ""
        if ocr_context is None:
            ocr_block = "Structured OCR evidence for the document:\n- Not included in this pass.\n"
        else:
            ocr_block = (
                "Structured OCR evidence for the document:\n"
                f"{json.dumps(ocr_context, indent=2)}\n"
            )
        return (
            f"Active workbook sheet: {sheet_name}\n\n"
            f"{self._workflow_step_context(session_state=session_state, sheet_name=sheet_name)}\n\n"
            f"{self._context_stage_instructions(context_stage)}\n\n"
            f"{sheet_reference_for_prompt(sheet_name)}\n\n"
            f"{sheet_specific_prompt_guidance(sheet_name)}\n"
            f"{workbook_context}\n\n"
            f"{scaffold_block}"
            f"{data_input_block}"
            f"{transactions_tool_block}"
            "Preloaded template data already present in this workbook before any agent writes:\n"
            f"{json.dumps(preloaded_template_context, indent=2)}\n\n"
            "Retained planner answers for this sheet:\n"
            f"{json.dumps(explicit_answers_payload, indent=2)}\n\n"
            "Retained agent decisions for this sheet:\n"
            f"{json.dumps(agent_answers_payload, indent=2)}\n\n"
            "Retained decisions are the only run-to-run memory carried into this pass. Agent-sourced decisions can come from a prior raw PDF re-review or an explicit user delegation. If they are present, resolve those targets yourself and avoid asking the same question again unless the sheet still cannot proceed safely.\n\n"
            "Prior mapped assignments already written for this sheet:\n"
            f"{json.dumps(prior_assignments, indent=2)}\n\n"
            f"{ocr_block}"
        )

    def _build_raw_pdf_rereview_prompt(
        self,
        *,
        session_state: EntrySessionState,
        ocr_results: list[PageOcrResult],
        sheet_name: str,
        question: AgentQuestion,
    ) -> str:
        explicit_answers_payload = [
            answer.model_dump(mode="json")
            for answer in self._latest_answers_for_sheet(
                session_state,
                sheet_name=sheet_name,
                sources={"option", "free_text"},
            )
        ]
        workbook_context = read_sheet_context(
            session_state.workbook_path,
            sheet_name,
            touched_cells=touched_cells_for_assignments(session_state.mapped_assignments, sheet_name),
        )
        scaffold_context = read_sheet_scaffold_context(session_state.workbook_path, sheet_name)
        data_input_context = self._data_input_context(
            session_state=session_state,
            sheet_name=sheet_name,
        )
        preloaded_template_context = read_preloaded_template_context(session_state.workbook_path)
        raw_ocr_payload = [
            {
                "page_number": result.page_number,
                "raw_text": result.raw_text,
                "source_snippets": result.source_snippets,
            }
            for result in ocr_results
        ]
        data_input_block = ""
        if data_input_context is not None:
            data_input_block = f"Populated Data Input sheet context already written into the workbook:\n{data_input_context}\n\n"
        scaffold_block = ""
        if scaffold_context is not None:
            scaffold_block = f"{scaffold_context}\n\n"
        return (
            f"Active workbook sheet: {sheet_name}\n\n"
            f"{self._workflow_step_context(session_state=session_state, sheet_name=sheet_name)}\n\n"
            "Context stage: raw_pdf_review\n"
            "This is the final fallback. Use workbook context first, then the raw OCR text below only to answer the pending question.\n\n"
            f"{sheet_reference_for_prompt(sheet_name)}\n\n"
            f"{sheet_specific_prompt_guidance(sheet_name)}\n"
            f"{workbook_context}\n\n"
            f"{scaffold_block}"
            f"{data_input_block}"
            "Preloaded template data already present in this workbook before any agent writes:\n"
            f"{json.dumps(preloaded_template_context, indent=2)}\n\n"
            "Prior explicit user answers for this sheet:\n"
            f"{json.dumps(explicit_answers_payload, indent=2)}\n\n"
            "Pending question JSON:\n"
            f"{json.dumps(question.model_dump(mode='json'), indent=2)}\n\n"
            "Raw OCR text and source snippets for the full document:\n"
            f"{json.dumps(raw_ocr_payload, indent=2)}\n\n"
            "If the raw OCR gives a defensible answer to the pending question, return answer_found=true and provide the answer text exactly as it should be used. If not, return answer_found=false."
        )

    def _review_question_against_raw_pdf(
        self,
        *,
        session_state: EntrySessionState,
        ocr_results: list[PageOcrResult],
        sheet_name: str,
        question: AgentQuestion,
        retry_notifier: RetryNotifier | None = None,
        usage_notifier: UsageNotifier | None = None,
        progress_notifier: ProgressNotifier | None = None,
    ) -> AgentAnswer | None:
        review = self._invoke_structured_with_retries(
            operation_name=f"Raw PDF re-review for {sheet_name}",
            schema=RawPdfQuestionReview,
            messages=[
                SystemMessage(content=RAW_PDF_REREVIEW_SYSTEM_PROMPT),
                HumanMessage(
                    content=self._build_raw_pdf_rereview_prompt(
                        session_state=session_state,
                        ocr_results=ocr_results,
                        sheet_name=sheet_name,
                        question=question,
                    )
                ),
            ],
            retry_notifier=retry_notifier,
            usage_notifier=usage_notifier,
            progress_notifier=progress_notifier,
        )
        if not review.answer_found or review.answer is None or not review.answer.strip():
            return None
        return answer_from_question(
            question,
            answer=review.answer.strip(),
            source="raw_pdf_review",
        )

    @staticmethod
    def _sheet_has_raw_pdf_rereview_answer(
        session_state: EntrySessionState,
        sheet_name: str,
    ) -> bool:
        return any(
            answer.sheet_name == sheet_name and answer.source == "raw_pdf_review"
            for answer in session_state.user_answers
        )

    def _run_sheet_pass(
        self,
        *,
        session_state: EntrySessionState,
        ocr_results: list[PageOcrResult],
        sheet_name: str,
        context_stage: Literal["workbook_only", "ocr_results"],
        retry_notifier: RetryNotifier | None,
        usage_notifier: UsageNotifier | None,
        progress_notifier: ProgressNotifier | None,
    ) -> SheetEntryResult:
        sheet_tools = self._sheet_tools(
            session_state=session_state,
            sheet_name=sheet_name,
        )
        graph_state = self._graph.invoke(
            {
                "sheet_name": sheet_name,
                "prompt": self._build_prompt(
                    session_state=session_state,
                    ocr_results=ocr_results,
                    sheet_name=sheet_name,
                    context_stage=context_stage,
                ),
                "retry_notifier": retry_notifier,
                "usage_notifier": usage_notifier,
                "progress_notifier": progress_notifier,
                "tools": sheet_tools,
                "result": None,
            }
        )
        result = graph_state["result"]
        if result is None:
            raise RuntimeError(f"Template entry graph did not produce a result for {sheet_name}.")
        result.sheet_name = sheet_name
        return result

    def _initial_context_stage(
        self,
        *,
        session_state: EntrySessionState,
        sheet_name: str,
    ) -> Literal["workbook_only", "ocr_results"]:
        if sheet_name == "Data Input":
            return "ocr_results"
        if sheet_has_populated_writable_cells(session_state.workbook_path, "Data Input"):
            return "workbook_only"
        return "ocr_results"

    def _should_escalate_to_ocr(
        self,
        *,
        context_stage: Literal["workbook_only", "ocr_results"],
        result: SheetEntryResult,
    ) -> bool:
        if context_stage != "workbook_only":
            return False
        return (
            result.question is not None
            or bool(result.unresolved_supported_targets)
            or self._mapped_item_count(result) == 0
        )

    def advance(
        self,
        session_state: EntrySessionState,
        ocr_results: list[PageOcrResult],
        retry_notifier: RetryNotifier | None = None,
        usage_notifier: UsageNotifier | None = None,
        progress_notifier: ProgressNotifier | None = None,
    ) -> EntrySessionState:
        return self._advance(
            session_state,
            ocr_results,
            allow_raw_pdf_rereview=True,
            retry_notifier=retry_notifier,
            usage_notifier=usage_notifier,
            progress_notifier=progress_notifier,
        )

    def _advance(
        self,
        session_state: EntrySessionState,
        ocr_results: list[PageOcrResult],
        *,
        allow_raw_pdf_rereview: bool,
        retry_notifier: RetryNotifier | None,
        usage_notifier: UsageNotifier | None,
        progress_notifier: ProgressNotifier | None,
    ) -> EntrySessionState:
        if session_state.current_sheet_index >= len(session_state.sheet_order):
            session_state.pending_question = None
            session_state.completed = True
            session_state.coverage_summary = coverage_summary_for_state(session_state)
            return session_state

        if session_state.current_sheet_index == 0:
            session_state.sheet_order = prioritize_data_input_sheet(session_state.sheet_order)

        sheet_name = session_state.sheet_order[session_state.current_sheet_index]
        if not ALLOWED_WRITE_CELLS_BY_SHEET.get(sheet_name):
            skip_message = "No writable targets configured for this sheet."
            if sheet_has_preloaded_template_data(session_state.workbook_path, sheet_name):
                skip_message = (
                    "No writable targets configured for this sheet. The locked template already "
                    "pre-populates this sheet with starter data."
                )
            _upsert_sheet_summary(
                session_state,
                SheetEntrySummary(
                    sheet_name=sheet_name,
                    status="skipped",
                    message=skip_message,
                )
            )
            session_state.current_sheet_index += 1
            session_state.coverage_summary = coverage_summary_for_state(session_state)
            session_state.completed = session_state.current_sheet_index >= len(session_state.sheet_order)
            return session_state

        context_stage = self._initial_context_stage(
            session_state=session_state,
            sheet_name=sheet_name,
        )
        result = self._run_sheet_pass(
            session_state=session_state,
            ocr_results=ocr_results,
            sheet_name=sheet_name,
            context_stage=context_stage,
            retry_notifier=retry_notifier,
            usage_notifier=usage_notifier,
            progress_notifier=progress_notifier,
        )
        if self._should_escalate_to_ocr(context_stage=context_stage, result=result):
            result = self._run_sheet_pass(
                session_state=session_state,
                ocr_results=ocr_results,
                sheet_name=sheet_name,
                context_stage="ocr_results",
                retry_notifier=retry_notifier,
                usage_notifier=usage_notifier,
                progress_notifier=progress_notifier,
            )
        if result.question is not None and allow_raw_pdf_rereview:
            reviewed_answer = self._review_question_against_raw_pdf(
                session_state=session_state,
                ocr_results=ocr_results,
                sheet_name=sheet_name,
                question=result.question,
                retry_notifier=retry_notifier,
                usage_notifier=usage_notifier,
                progress_notifier=progress_notifier,
            )
            if reviewed_answer is not None:
                session_state.user_answers.append(reviewed_answer)
                result = self._run_sheet_pass(
                    session_state=session_state,
                    ocr_results=ocr_results,
                    sheet_name=sheet_name,
                    context_stage="ocr_results",
                    retry_notifier=retry_notifier,
                    usage_notifier=usage_notifier,
                    progress_notifier=progress_notifier,
                )
            if result.question is not None:
                result.question.pdf_rereviewed = True

        if (
            result.question is not None
            and self._sheet_has_raw_pdf_rereview_answer(session_state, sheet_name)
        ):
            result.question.pdf_rereviewed = True

        if result.question is not None:
            session_state.pending_question = result.question
            session_state.questions_asked.append(result.question)
            _upsert_sheet_summary(
                session_state,
                SheetEntrySummary(
                    sheet_name=sheet_name,
                    status="needs_input",
                    mapped_count=(
                        len(result.mapped_fields)
                        + len(result.expenses)
                        + len(result.accounts)
                        + len(result.holdings)
                    ),
                    unresolved_count=len(result.unresolved_supported_targets),
                    message=result.question.prompt,
                )
            )
            session_state.coverage_summary = coverage_summary_for_state(session_state)
            return session_state

        session_state.pending_question = None
        session_state.sheet_results.append(result)
        _upsert_sheet_summary(
            session_state,
            SheetEntrySummary(
                sheet_name=sheet_name,
                status="completed",
                mapped_count=(
                    len(result.mapped_fields)
                    + len(result.expenses)
                    + len(result.accounts)
                    + len(result.holdings)
                ),
                unresolved_count=len(result.unresolved_supported_targets),
                message="Sheet entry completed.",
            )
        )
        session_state.current_sheet_index += 1
        session_state.coverage_summary = coverage_summary_for_state(session_state)
        session_state.completed = session_state.current_sheet_index >= len(session_state.sheet_order)
        return session_state


def default_sheet_order() -> list[str]:
    return prioritize_data_input_sheet(
        [sheet_name for sheet_name in TEMPLATE_SHEET_ORDER if sheet_name != TRANSACTIONS_SHEET_NAME]
    )


def answer_from_question(
    question: AgentQuestion,
    *,
    answer: str,
    source: str,
) -> AgentAnswer:
    if source == "free_text":
        normalized_source = "free_text"
    elif source == "agent":
        normalized_source = "agent"
    elif source == "raw_pdf_review":
        normalized_source = "raw_pdf_review"
    else:
        normalized_source = "option"
    return AgentAnswer(
        question_id=question.id,
        sheet_name=question.sheet_name,
        answer=answer,
        source=normalized_source,
        affected_targets=question.affected_targets,
    )


def fallback_question(
    *,
    sheet_name: str,
    prompt: str,
    rationale: str,
    affected_targets: list[str],
    options: list[tuple[str, str, str]],
) -> AgentQuestion:
    return AgentQuestion(
        id=f"{sheet_name.lower().replace(' ', '_')}-question",
        sheet_name=sheet_name,
        prompt=prompt,
        rationale=rationale,
        affected_targets=affected_targets,
        options=[
            QuestionOption(label=label, value=value, description=description)
            for label, value, description in options
        ],
    )
