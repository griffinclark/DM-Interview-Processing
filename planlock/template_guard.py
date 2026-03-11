from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from planlock.models import TemplateDriftCheckResult


IGNORED_PACKAGE_PARTS = {
    "docProps/core.xml",
    "xl/calcChain.xml",
}


@dataclass(frozen=True)
class CellSnapshot:
    value: str | None
    formula: str | None
    style_id: int
    data_type: str


@dataclass(frozen=True)
class SheetSnapshot:
    title: str
    state: str
    cells: dict[str, CellSnapshot]
    merged_ranges: tuple[str, ...]
    data_validations: tuple[str, ...]
    conditional_formatting: tuple[str, ...]
    row_dimensions: tuple[str, ...]
    column_dimensions: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookSnapshot:
    package_parts: tuple[str, ...]
    sheet_names: tuple[str, ...]
    sheet_snapshots: dict[str, SheetSnapshot]
    defined_names: tuple[str, ...]


def _package_parts(path: Path) -> tuple[str, ...]:
    with ZipFile(path) as archive:
        return tuple(sorted(name for name in archive.namelist() if name not in IGNORED_PACKAGE_PARTS))


def _defined_names(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path)
    return tuple(
        sorted(
            f"{name}:{definition.attr_text}"
            for name, definition in workbook.defined_names.items()
        )
    )


def _sheet_snapshot(path: Path) -> dict[str, SheetSnapshot]:
    workbook = load_workbook(path)
    snapshots: dict[str, SheetSnapshot] = {}
    for sheet in workbook.worksheets:
        cells: dict[str, CellSnapshot] = {}
        for coordinate, cell in sheet._cells.items():
            if isinstance(coordinate, tuple):
                row, column = coordinate
                cell_ref = f"{get_column_letter(column)}{row}"
            else:
                cell_ref = coordinate
            value = cell.value
            formula = value if isinstance(value, str) and value.startswith("=") else None
            normalized_value = None if formula else (None if value is None else str(value))
            cells[cell_ref] = CellSnapshot(
                value=normalized_value,
                formula=formula,
                style_id=cell.style_id,
                data_type=cell.data_type,
            )

        data_validations = ()
        if sheet.data_validations is not None:
            data_validations = tuple(
                sorted(
                    f"{validation.type}:{validation.sqref}:{validation.formula1}:{validation.formula2}"
                    for validation in sheet.data_validations.dataValidation
                )
            )

        conditional_formatting = tuple(
            sorted(
                f"{str(sqref)}:{len(rules)}"
                for sqref, rules in sheet.conditional_formatting._cf_rules.items()
            )
        )

        row_dimensions = tuple(
            sorted(
                f"{key}:{dimension.hidden}:{dimension.outlineLevel}:{dimension.height}"
                for key, dimension in sheet.row_dimensions.items()
                if dimension.hidden or dimension.outlineLevel or dimension.height
            )
        )
        column_dimensions = tuple(
            sorted(
                f"{key}:{dimension.hidden}:{dimension.outline_level}:{dimension.width}"
                for key, dimension in sheet.column_dimensions.items()
                if dimension.hidden or dimension.outline_level or dimension.width
            )
        )

        snapshots[sheet.title] = SheetSnapshot(
            title=sheet.title,
            state=sheet.sheet_state,
            cells=cells,
            merged_ranges=tuple(sorted(str(cell_range) for cell_range in sheet.merged_cells.ranges)),
            data_validations=data_validations,
            conditional_formatting=conditional_formatting,
            row_dimensions=row_dimensions,
            column_dimensions=column_dimensions,
        )
    return snapshots


def create_workbook_snapshot(path: Path) -> WorkbookSnapshot:
    workbook = load_workbook(path)
    return WorkbookSnapshot(
        package_parts=_package_parts(path),
        sheet_names=tuple(workbook.sheetnames),
        sheet_snapshots=_sheet_snapshot(path),
        defined_names=_defined_names(path),
    )


def check_for_drift(
    template_path: Path,
    output_path: Path,
    allowed_cells_by_sheet: dict[str, set[str]],
) -> TemplateDriftCheckResult:
    template = create_workbook_snapshot(template_path)
    output = create_workbook_snapshot(output_path)
    violations: list[str] = []

    if template.package_parts != output.package_parts:
        violations.append("Workbook package parts changed outside allowed differences.")
    if template.sheet_names != output.sheet_names:
        violations.append("Workbook sheet names or order changed.")
    if template.defined_names != output.defined_names:
        violations.append("Workbook defined names changed.")

    for sheet_name in template.sheet_names:
        template_sheet = template.sheet_snapshots[sheet_name]
        output_sheet = output.sheet_snapshots.get(sheet_name)
        if output_sheet is None:
            violations.append(f"Missing sheet in output workbook: {sheet_name}")
            continue

        if template_sheet.state != output_sheet.state:
            violations.append(f"Sheet visibility changed for {sheet_name}.")
        if template_sheet.merged_ranges != output_sheet.merged_ranges:
            violations.append(f"Merged ranges changed for {sheet_name}.")
        if template_sheet.data_validations != output_sheet.data_validations:
            violations.append(f"Data validations changed for {sheet_name}.")
        if template_sheet.conditional_formatting != output_sheet.conditional_formatting:
            violations.append(f"Conditional formatting changed for {sheet_name}.")
        if template_sheet.row_dimensions != output_sheet.row_dimensions:
            violations.append(f"Row dimensions changed for {sheet_name}.")
        if template_sheet.column_dimensions != output_sheet.column_dimensions:
            violations.append(f"Column dimensions changed for {sheet_name}.")

        allowed_cells = allowed_cells_by_sheet.get(sheet_name, set())
        all_cells = set(template_sheet.cells) | set(output_sheet.cells)
        for cell in sorted(all_cells):
            template_cell = template_sheet.cells.get(cell)
            output_cell = output_sheet.cells.get(cell)
            if cell in allowed_cells:
                if template_cell and output_cell and template_cell.style_id != output_cell.style_id:
                    violations.append(f"Style drift in allowed cell {sheet_name}!{cell}.")
                continue

            if template_cell != output_cell:
                violations.append(f"Non-whitelisted cell changed: {sheet_name}!{cell}.")
                if len(violations) >= 25:
                    return TemplateDriftCheckResult(passed=False, violations=violations)

    return TemplateDriftCheckResult(passed=not violations, violations=violations)
