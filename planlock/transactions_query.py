from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
import json
from pathlib import Path
import re
import sqlite3
from typing import Any

from langchain_core.tools import BaseTool, StructuredTool
from openpyxl import load_workbook
from pydantic import BaseModel, Field


TRANSACTIONS_SHEET_NAME = "Transactions Raw"
TRANSACTIONS_TABLE_NAME = "transactions_raw"
TRANSACTIONS_CELLS_TABLE_NAME = "transactions_raw_cells"
TRANSACTIONS_QUERY_MAX_ROWS = 200
_READ_ONLY_SQL_PATTERN = re.compile(r"^\s*(select|with|pragma\s+table_info\b)", re.IGNORECASE)
_DISALLOWED_SQL_PATTERN = re.compile(
    r"\b(insert|update|delete|drop|alter|create|attach|detach|replace|truncate|vacuum|reindex)\b",
    re.IGNORECASE,
)
_COLUMN_SPECS = [
    ("A", "account"),
    ("B", "date_posted"),
    ("C", "amount"),
    ("D", "merchant"),
    ("E", "description"),
    ("F", "default_category"),
]


class TransactionQueryInput(BaseModel):
    sql: str = Field(
        description=(
            "A read-only SQLite query against transactions_raw or transactions_raw_cells. "
            "Use SELECT, WITH, or PRAGMA table_info(...)."
        )
    )


def transactions_query_schema_reference() -> str:
    return (
        "Transactions Raw query tool schema:\n"
        "- Table `transactions_raw`: `sheet_name` TEXT, `row_number` INTEGER, `account` TEXT, "
        "`date_posted` TEXT (ISO-8601), `amount` REAL, `merchant` TEXT, `description` TEXT, "
        "`default_category` TEXT.\n"
        "- Table `transactions_raw_cells`: `sheet_name` TEXT, `row_number` INTEGER, "
        "`column_letter` TEXT, `column_name` TEXT, `value_text` TEXT.\n"
        "- Excel column mapping: A=`account`, B=`date_posted`, C=`amount`, D=`merchant`, "
        "E=`description`, F=`default_category`.\n"
        "- Negative `amount` values are outflows/spend. Positive `amount` values are inflows, "
        "refunds, transfers, or credits.\n"
        f"- Tool results are capped at {TRANSACTIONS_QUERY_MAX_ROWS} rows per query."
    )


def has_transaction_data(workbook_path: Path) -> bool:
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        if TRANSACTIONS_SHEET_NAME not in workbook.sheetnames:
            return False
        sheet = workbook[TRANSACTIONS_SHEET_NAME]
        populated_rows = 0
        for values in sheet.iter_rows(min_row=1, max_col=len(_COLUMN_SPECS), values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            populated_rows += 1
            if populated_rows > 1:
                return True
        return False
    finally:
        workbook.close()


def _normalize_scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _serialize_result_value(value: Any) -> Any:
    normalized = _normalize_scalar(value)
    if isinstance(normalized, (str, int, float)) or normalized is None:
        return normalized
    return str(normalized)


def _validate_read_only_sql(sql: str) -> str:
    normalized = sql.strip()
    if not normalized:
        raise ValueError("SQL query cannot be empty.")
    if _DISALLOWED_SQL_PATTERN.search(normalized):
        raise ValueError("Only read-only SELECT, WITH, or PRAGMA table_info queries are allowed.")
    if not _READ_ONLY_SQL_PATTERN.match(normalized):
        raise ValueError("Only read-only SELECT, WITH, or PRAGMA table_info queries are allowed.")
    return normalized


@dataclass
class TransactionQueryService:
    workbook_path: Path
    max_rows: int = TRANSACTIONS_QUERY_MAX_ROWS
    _connection: sqlite3.Connection | None = field(default=None, init=False, repr=False)

    def query(self, sql: str) -> dict[str, object]:
        normalized_sql = _validate_read_only_sql(sql)
        connection = self._connection_or_build()
        cursor = connection.execute(normalized_sql)
        rows = cursor.fetchmany(self.max_rows + 1)
        truncated = len(rows) > self.max_rows
        visible_rows = rows[: self.max_rows]
        columns = [description[0] for description in cursor.description or []]
        return {
            "status": "ok",
            "sql": normalized_sql,
            "columns": columns,
            "row_count": len(visible_rows),
            "truncated": truncated,
            "rows": [
                {
                    column: _serialize_result_value(row[column])
                    for column in columns
                }
                for row in visible_rows
            ],
        }

    def _connection_or_build(self) -> sqlite3.Connection:
        if self._connection is None:
            self._connection = self._build_connection()
        return self._connection

    def _build_connection(self) -> sqlite3.Connection:
        workbook = load_workbook(self.workbook_path, data_only=False, read_only=True)
        if TRANSACTIONS_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Workbook does not contain a '{TRANSACTIONS_SHEET_NAME}' sheet.")

        connection = sqlite3.connect(":memory:")
        connection.row_factory = sqlite3.Row
        connection.execute(
            f"""
            CREATE TABLE {TRANSACTIONS_TABLE_NAME} (
                sheet_name TEXT NOT NULL,
                row_number INTEGER NOT NULL,
                account TEXT,
                date_posted TEXT,
                amount REAL,
                merchant TEXT,
                description TEXT,
                default_category TEXT,
                PRIMARY KEY (sheet_name, row_number)
            )
            """
        )
        connection.execute(
            f"""
            CREATE TABLE {TRANSACTIONS_CELLS_TABLE_NAME} (
                sheet_name TEXT NOT NULL,
                row_number INTEGER NOT NULL,
                column_letter TEXT NOT NULL,
                column_name TEXT NOT NULL,
                value_text TEXT
            )
            """
        )

        sheet = workbook[TRANSACTIONS_SHEET_NAME]
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(_COLUMN_SPECS), values_only=True),
            start=2,
        ):
            if not any(value not in (None, "") for value in values):
                continue

            normalized_values = [_normalize_scalar(value) for value in values]
            amount_value = normalized_values[2]
            if amount_value not in (None, "") and not isinstance(amount_value, (int, float)):
                try:
                    amount_value = float(str(amount_value))
                except ValueError:
                    amount_value = None

            connection.execute(
                f"""
                INSERT INTO {TRANSACTIONS_TABLE_NAME} (
                    sheet_name,
                    row_number,
                    account,
                    date_posted,
                    amount,
                    merchant,
                    description,
                    default_category
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    TRANSACTIONS_SHEET_NAME,
                    row_number,
                    normalized_values[0],
                    normalized_values[1],
                    amount_value,
                    normalized_values[3],
                    normalized_values[4],
                    normalized_values[5],
                ),
            )

            for (column_letter, column_name), value in zip(_COLUMN_SPECS, normalized_values, strict=False):
                if value in (None, ""):
                    continue
                connection.execute(
                    f"""
                    INSERT INTO {TRANSACTIONS_CELLS_TABLE_NAME} (
                        sheet_name,
                        row_number,
                        column_letter,
                        column_name,
                        value_text
                    ) VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        TRANSACTIONS_SHEET_NAME,
                        row_number,
                        column_letter,
                        column_name,
                        str(value),
                    ),
                )

        connection.execute(
            f"CREATE INDEX idx_{TRANSACTIONS_TABLE_NAME}_default_category "
            f"ON {TRANSACTIONS_TABLE_NAME} (default_category)"
        )
        connection.execute(
            f"CREATE INDEX idx_{TRANSACTIONS_TABLE_NAME}_merchant "
            f"ON {TRANSACTIONS_TABLE_NAME} (merchant)"
        )
        connection.execute(
            f"CREATE INDEX idx_{TRANSACTIONS_CELLS_TABLE_NAME}_row_col "
            f"ON {TRANSACTIONS_CELLS_TABLE_NAME} (row_number, column_letter)"
        )
        return connection


def build_query_transactions_tool(
    workbook_path: Path,
    *,
    max_rows: int = TRANSACTIONS_QUERY_MAX_ROWS,
) -> BaseTool:
    service = TransactionQueryService(workbook_path=workbook_path, max_rows=max_rows)

    def _query_transactions(sql: str) -> str:
        try:
            payload = service.query(sql)
        except Exception as exc:  # noqa: BLE001
            payload = {
                "status": "error",
                "sql": sql,
                "error": str(exc),
                "schema": transactions_query_schema_reference(),
            }
        return json.dumps(payload, ensure_ascii=True)

    return StructuredTool.from_function(
        name="query_transactions",
        description=(
            "Run a read-only SQLite query against the workbook's Transactions Raw sheet. "
            "Use `transactions_raw` for typed transaction rows and `transactions_raw_cells` "
            "for explicit row/column provenance. "
            f"{transactions_query_schema_reference()}"
        ),
        func=_query_transactions,
        args_schema=TransactionQueryInput,
    )
